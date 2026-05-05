import openpyxl
import os
import csv
import zipfile
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from copy import copy
from datetime import datetime

"""
Automated Payroll Report Formatter
----------------------------------
This script automates the transformation of raw workforce management exports 
into standardized, audit-ready financial reports. 

Features:
- Multi-format ingestion (XLSX/CSV) with automated type conversion.
- Logic-based row filtering (exclusion of non-billable entries).
- Dynamic date range extraction for header generation.
- Precise UI-driven metadata collection using Tkinter.
- Low-level XML patching to achieve layout standards beyond standard API capabilities.
"""

def load_source_file(file_path):
    """
    Ingests the source file. If the file is not a native Workbook (e.g., CSV), 
    it parses the raw data and builds a temporary memory-mapped workbook.
    """
    try:
        return openpyxl.load_workbook(file_path)
    except Exception:
        wb = openpyxl.Workbook()
        ws = wb.active
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader, 1):
                for col_idx, value in enumerate(row, 1):
                    try:
                        # Automated type-casting for numerical analysis
                        if "." in value:
                            ws.cell(row=row_idx, column=col_idx).value = float(value)
                        else:
                            ws.cell(row=row_idx, column=col_idx).value = int(value)
                    except ValueError:
                        ws.cell(row=row_idx, column=col_idx).value = value
        return wb

def get_report_metadata(root):
    """
    Launches a modal dialog to collect necessary reporting metadata.
    """
    dialog = tk.Toplevel(root)
    dialog.title("Report Configuration")
    dialog.geometry("350x120")
    dialog.resizable(False, False)
    dialog.grab_set()

    tk.Label(dialog, text="Department/Unit:").grid(row=0, column=0, padx=10, pady=10, sticky='w')
    unit_var = tk.StringVar()
    unit_entry = tk.Entry(dialog, textvariable=unit_var, width=30)
    unit_entry.grid(row=0, column=1, padx=10, pady=10)

    result = {}

    def on_ok():
        result['unit'] = unit_var.get().strip()
        dialog.destroy()

    def on_cancel():
        result['cancelled'] = True
        dialog.destroy()

    tk.Button(dialog, text="Process", width=10, command=on_ok).grid(row=1, column=0, padx=10, pady=10)
    tk.Button(dialog, text="Cancel", width=10, command=on_cancel).grid(row=1, column=1, padx=10, pady=10)

    unit_entry.focus()
    dialog.wait_window()
    return result


def process_report():
    """
    Main execution pipeline for data transformation and styling.
    """
    root = tk.Tk()
    root.withdraw()

    info = get_report_metadata(root)
    if info.get('cancelled'):
        root.destroy()
        return

    report_unit = info.get('unit', 'General')
    report_category = 'Operational' # Standardized category

    file_path = filedialog.askopenfilename(
        title="Select Source Export File",
        filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files", "*.*")]
    )
    if not file_path:
        root.destroy()
        return

    try:
        wb = load_source_file(file_path)
        ws = wb.active


        # --- DATA CLEANING: Remove Non-Billable/On-Call Rows ---
        # Logic: Scans for specific status codes and flags rows for deletion.
        # Deletion is handled in reverse order to maintain index integrity.
        rows_to_delete = []
        for row in ws.iter_rows():
            exclude_row = False
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "EXCLUDE-NON-PAY" in cell.value.upper():
                    exclude_row = True
                    break
            # Specific code filtering (e.g., placeholder for administrative codes)
            if len(row) > 3 and row[3].value and isinstance(row[3].value, str) and row[3].value.startswith("NC-"):
                exclude_row = True
            
            if exclude_row:
                rows_to_delete.append(row[0].row)

        for r in reversed(rows_to_delete):
            ws.delete_rows(r)


        # --- DYNAMIC DATE EXTRACTION ---
        # Determines report period based on the data payload.
        min_date, max_date = None, None
        for row in ws.iter_rows(min_row=4, min_col=3, max_col=3):
            cell_val = row[0].value
            if isinstance(cell_val, datetime):
                dt = cell_val
            elif cell_val and isinstance(cell_val, str):
                try:
                    dt = datetime.strptime(cell_val.split(' ')[0], "%m/%d/%Y")
                except:
                    continue
            else:
                continue
            if not min_date or dt < min_date: min_date = dt
            if not max_date or dt > max_date: max_date = dt
        period_end_str = max_date.strftime('%m/%d/%y') if max_date else "N/A"
        process_date = datetime.today().strftime('%m/%d/%y')
        period_range = f"{min_date.strftime('%m/%d/%y')} - {period_end_str}" if min_date else "Unknown"

        # --- HEADER & BORDER STYLING ---
        # Adjusts cell positioning and applies high-visibility reporting styles.
        ws['A2'].value = ws['A1'].value # Shift primary identifier
        ws['A1'].value = None
        red_bold = Font(color="FFFF0000", bold=True, size=12, name='Calibri')

        ws['E1'].value = "ENTRY ID"
        ws['E1'].font = red_bold
        
        ws['F1'].value = f"TOTALS THROUGH {period_end_str}"
        ws.merge_cells('F1:G1')
        ws['F1'].font = red_bold
        ws['F1'].alignment = Alignment(horizontal='center')

        # Apply horizontal rule to header
        medium_border = Side(style='medium')
        for col in range(1, 8):
            ws.cell(1, col).border = Border(bottom=medium_border)

        # --- DATA MAPPING & ALIGNMENT ---
        # Maps raw columns to standardized financial templates and applies formatting.
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            if row[0].row >= 4:
                row[4].value = None # Clear unnecessary metadata
            
            # Map hours to target column
            if len(row) > 8:
                hours_val = row[8].value
                if hours_val:
                    row[5].value = hours_val
                    row[8].value = None
            
            # Clear trailing data columns
            for cell in row[7:]:
                cell.value = None


        # Style numerical data (2 decimal precision)
        thin_border = Side(style='thin')
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            hour_cell = row[5]
            if hour_cell.value is not None:
                hour_cell.alignment = Alignment(horizontal='right', vertical='center')
                hour_cell.number_format = '0.00'
            
            # Apply row separators for readability
            if row[0].value and not row[1].value:
                for cell in row[:7]:
                    cell.border = Border(bottom=thin_border)

        # --- LAYOUT & PRINT CONFIGURATION ---
        # Implements specific column widths and row heights to match legacy templates.
        widths = {'A': 0.85, 'B': 3.70, 'C': 10.20, 'D': 4.58, 'E': 22.05, 'F': 8.54, 'G': 38.25}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.row_dimensions[1].height = 16.56
        ws.sheet_format.defaultRowHeight = 15.75
        ws.sheet_view.view = 'pageLayout'

        # Global Header Injection (Navy Style)
        ws.oddHeader.center.text = (
            f"Automated Workforce Report\n"
            f"Processed: {process_date} ({period_range})\n"
            f"Category: {report_category} | Dept: {report_unit}"
        )
        ws.oddHeader.center.size = 12
        ws.oddHeader.center.font = "Calibri,Bold"
        ws.oddHeader.center.color = "002060"


        # --- SAVE & XML INJECTION ---
        # Standard save followed by a low-level XML patch to control layout attributes
        # not exposed by the high-level API (e.g., dyDescent for precise row rendering).
        output_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=f"Formatted_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        
        if output_path:
            wb.save(output_path)
            tmp_path = output_path + ".tmp"
            
            with zipfile.ZipFile(output_path, 'r') as zin:
                with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                    for item in zin.infolist():
                        data = zin.read(item.filename)
                        if item.filename == 'xl/worksheets/sheet1.xml':
                            data = data.decode().replace(
                                '<sheetFormatPr',
                                '<sheetFormatPr xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"'
                            ).replace(
                                'defaultRowHeight="15.75"',
                                'defaultRowHeight="15.75" x14ac:dyDescent="0.2"'
                            ).encode()
                        zout.writestr(item, data)
            
            os.replace(tmp_path, output_path)
            messagebox.showinfo("Success", "Report transformation complete.")

    except Exception as e:
        messagebox.showerror("Error", f"Processing Failed:\n{str(e)}")
    finally:
        root.destroy()


if __name__ == "__main__":
    process_report()